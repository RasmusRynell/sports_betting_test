import data_processing.new_data_handling as data_handling
import openpyxl
import Settings

games_obj = data_handling.games()


def check_or_create_initial(sheet):
    (start_row, start_column, row_index, column_index) = (0,0,0,0)
    if sheet.cell(row=1, column=1).value == None:
        sheet.cell(row=1, column=1).value = "Start rows:"
        sheet.cell(row=1, column=2).value = 5
        start_row = 5
    else:
        start_row = int(sheet.cell(row=1, column=2).value)

    if sheet.cell(row=2, column=1).value == None:
        sheet.cell(row=2, column=1).value = "Start columns:"
        sheet.cell(row=2, column=2).value = 2
        start_column = 2
    else:
        start_column = int(sheet.cell(row=2, column=2).value)

    if sheet.cell(row=1, column=4).value == None:
        sheet.cell(row=1, column=4).value = "Total rows:"
        sheet.cell(row=1, column=5).value = 0
        row_index = 0
    else:
        row_index = int(sheet.cell(row=1, column=5).value)

    if sheet.cell(row=2, column=4).value == None:
        sheet.cell(row=2, column=4).value = "Total columns:"
        sheet.cell(row=2, column=5).value = 0
        column_index = 0
    else:
        column_index = int(sheet.cell(row=2, column=5).value)
    
    
    return (start_row, start_column, row_index + start_row, column_index + start_column)


def get_headers(sheet, start_row, start_column):
    done = []
    i = 2
    while sheet.cell(row=start_row - 1, column=i).value != None:
        done.append(str(sheet.cell(row=start_row - 1, column=i).value).lower())
        i += 1
    return done


def write_in_sheet(sheet, value):
    (start_row, start_column, row_index, column_index) = check_or_create_initial(sheet)

    # Get headers
    headers = get_headers(sheet, start_row, start_column)

    new_columns = 0
    for h, v in value.items():
        if h.lower() in headers:
            sheet.cell(row=row_index, column=(start_column + headers.index(h.lower()))).value = v
        else:
            sheet.cell(row=start_row - 1, column=column_index).value = h
            sheet.cell(row=row_index, column=column_index).value = v
            column_index += 1

    sheet.cell(row=1, column=5).value = row_index - start_row + 1
    sheet.cell(row=2, column=5).value = column_index - start_column



def write_dict(data, h, v, book):
    if (h == "basic_data"):
        write_in_sheet(book["basic_data"], v)
    elif (h == "predictions"):
        for header, value in v.items():
            if header not in book.sheetnames:
                book.create_sheet(header)
            write_in_sheet(book[header], value)



def save_data(data, path = './excel/main.xlsx'):
    book = openpyxl.load_workbook(filename=path)

    for header, value in data.items():
        write_dict(data, header, value, book)

    update_shots(path)

    book.save(path)



def get_bets(book):
    sheet = book['basic_data']
    (start_row, start_column, row_index, column_index) = check_or_create_initial(sheet)

    keys = []
    for key in range(start_column, column_index):
        keys.append(str(sheet.cell(row=start_row - 1, column=key).value))

    total = []
    for x in range(start_row, row_index):
        bet = {}
        for y in range(start_column, column_index):
            bet[keys[y-start_column]] = str(sheet.cell(row=x, column=y).value)
        total.append(bet)

    return total


def get_goals_done(bet):

    player_id = games_obj.get_player_id(bet["player_name"])
    ply_team_id = games_obj.get_team_id(bet["player_team"])
    opp_team_id =games_obj.get_team_id(bet["opp_team"])
    ids = (player_id, ply_team_id, opp_team_id)

    game_id = games_obj.get_game_from_date(player_id, bet["date"], ply_team_id, opp_team_id)
    
    answer = {
        "is_home_game": "-",
        "shots_this_game_O3.5": "-",
        "shots_this_game_O2.5": "-",
        "shots_this_game_O1.5": "-",
        "shots_this_game_U1.5": "-",
        "shots_this_game_U2.5": "-",
        "shots_this_game_U3.5": "-",
        "shots_this_game_total": "-"
    }

    answer.update(games_obj.calculate_answers(game_id, ids))

    return answer


def update_shots(path = './excel/main.xlsx'):
    book = openpyxl.load_workbook(filename=path)
    
    bets = get_bets(book)
    bets = [bet for bet in bets if bet != None]

    for bet in bets:
        bet.update(get_goals_done(bet))

    if "goals" not in book.sheetnames:
        book.create_sheet("goals")
    goals_sheet = book["goals"]
    (start_row, start_column, row_index, column_index) = check_or_create_initial(goals_sheet)
    row_index = start_row + len(bets)
    column_index = start_column + len(bets[0])

    header = list(bets[0].keys())
    for key in range(start_column, column_index):
        goals_sheet.cell(row=start_row - 1, column=key).value = header[key-start_column]


    for x in range(start_row, row_index):
        values = list(bets[x - start_row].values())
        for y in range(start_column, column_index):
            goals_sheet.cell(row=x, column=y).value = values[y-start_column]

    book.save(path)